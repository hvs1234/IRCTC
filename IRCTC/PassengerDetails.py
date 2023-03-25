#Libraries
from tkinter import *
from tkinter import messagebox
import openpyxl ; from openpyxl import Workbook
import pathlib

#Application Setup
root = Toplevel()
root.title("Passengers Detail")
root.geometry("600x400+340+140")
root.resizable(False,False)
root.configure(bg="light blue")

#Functions
file = pathlib.Path("Passenger Data.xlsx")
if(file.exists()): pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Passenger Name"
    sheet['B1'] = "Age"
    sheet['C1'] = "Phoneno"
    sheet['D1'] = "Gender"
    sheet['E1'] = "Book Date"
    file.save("Passenger Data.xlsx")

def submit():
    a = name_entry.get()
    b = age_entry.get()
    c = phone_entry.get()
    d = gender_entry.get()
    e = book_date_entry.get()

    try:
        if(b=='' and c==''): pass
        else:
            int(b); int(c)
    except ValueError:
        messagebox.showerror("Invalid Age","Please give your correct age and phone number")

    if(a=='' and (b=='0' or b=='') and (c=='0' or c=='') and d=='' and 
    (e=='dd/mm/yyyy' or e=='')):
        messagebox.showwarning("Blank Input","Please fill your details completely")
    if(a=='' or (b=='0' or b=='') or (c=='0' or c=='') or d=='' or 
    (e=='dd/mm/yyyy' or e=='')):
        messagebox.showwarning("Blank Input","Please fill your details completely")
    else: 
        file = openpyxl.load_workbook("Passenger Data.xlsx")
        sheet = file.active
        sheet.cell(column=1 , row = sheet.max_row+1 , value = a)
        sheet.cell(column=2 , row = sheet.max_row , value = b)
        sheet.cell(column=3 , row = sheet.max_row , value = c)
        sheet.cell(column=4 , row = sheet.max_row , value = d)
        sheet.cell(column=5 , row = sheet.max_row , value = e)
        file.save("Passenger Data.xlsx")
        messagebox.showinfo("Success","You have been booked successfully")

#Applciation Creation
img1 = PhotoImage(file="E:\\build app with python\\IRCTC\\passenger logo.png")
root.iconphoto(False,img1)

header = Label(root,text="🚊  Booking Details  🚊",font="Algerian 15 bold",bg="black",fg="aqua",width=300,height=2) ; header.pack()

name = Label(root,text="👤  Enter your name:",bg="light blue",fg="black",font="Gabriola 20 bold")
name.place(x=20,y=70)
age = Label(root,text="👥  Enter your age:",bg="light blue",fg="black",font="Gabriola 20 bold")
age.place(x=20,y=130)
phone = Label(root,text="📱  Enter your phoneno:",bg="light blue",fg="black",font="Gabriola 20 bold")
phone.place(x=20,y=190)
gender= Label(root,text="🚻  Enter your gender:",bg="light blue",fg="black",font="Gabriola 20 bold")
gender.place(x=20,y=250)
book_date = Label(root,text="📆  Select booking date:",bg="light blue",fg="black",font="Gabriola 20 bold")
book_date.place(x=20,y=310)

name1 = StringVar()
age1 = IntVar()
phone1 = IntVar()
gender1 = StringVar()
book_date1 = StringVar()

name_entry = Entry(root,width=40,textvariable=name1,bg="light grey",fg="black",bd=4)
name_entry.place(x=300,y=88)

age_entry = Entry(root,width=40,textvariable=age1,bg="light grey",fg="black",bd=4)
age_entry.place(x=300,y=148)

phone_entry = Entry(root,width=40,textvariable=phone1,bg="light grey",fg="black",bd=4)
phone_entry.place(x=300,y=208)
gender_entry = Entry(root,width=40,textvariable=gender1,bg="light grey",fg="black",bd=4)
gender_entry.place(x=300,y=268)
book_date_entry = Entry(root,width=40,textvariable=book_date1,bg="light grey",fg="black",bd=4) ; book_date_entry.insert(0,'dd/mm/yyyy')
book_date_entry.place(x=300,y=328)

submit = Button(root,text="Submit",bg="light blue",activebackground="light blue",
fg="dark blue",activeforeground="dark blue",bd=0,command=submit)
submit.place(x=400,y=370)


root.mainloop()
